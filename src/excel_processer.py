import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def list_sheets(file_path):
    """列出 Excel 中的所有工作表名稱"""
    try:
        sheets = pd.ExcelFile(file_path).sheet_names
        return sheets
    except Exception as e:
        print(f"無法讀取 Excel 檔案: {e}")
        return []

def read_excel(file_path, sheet_name):
    """讀取指定的工作表並返回資料框"""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df
    except Exception as e:
        print(f"無法讀取 Excel 檔案: {e}")
        return None

def export_to_excel(results, output_path):
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        all_data = []

        for result in results:
            row_header = pd.DataFrame(
                [[result['Row'], 'A', 'B', 'C', 'D', '幾何平均數', '權重(W)', '一致性指標']],
                columns=['index', 'A', 'B', 'C', 'D', '幾何平均數', '權重(W)', '一致性指標']
            )
            all_data.append(row_header)

            df_matrix = pd.DataFrame(
                result['Matrix'],
                columns=['A', 'B', 'C', 'D'],
                index=['A', 'B', 'C', 'D']
            ).reset_index()
            
            df_matrix['幾何平均數'] = result['Geometric Mean']
            df_matrix['權重(W)'] = result['Weights']

            df_matrix.at[0, '一致性指標'] = '一致性指數'
            df_matrix.at[1, '一致性指標'] = result['CI']
            df_matrix.at[2, '一致性指標'] = '一致性比率'
            df_matrix.at[3, '一致性指標'] = result['CR']

            all_data.append(df_matrix)

            all_data.append(pd.DataFrame([[''] * len(df_matrix.columns)], columns=df_matrix.columns))

        final_data = pd.concat(all_data, ignore_index=True)

        final_data.to_excel(writer, sheet_name="Analysis_Results", index=False)
        
    workbook = load_workbook(output_path)
    sheet = workbook.active
    red_font = Font(color="FF0000")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=8, max_col=8):
        for cell in row:
            if isinstance(cell.value, float) and cell.value > 0.1:
                cell.font = red_font

    workbook.save(output_path)
